from django.db.models import Count, Q
from django.utils import timezone
from rest_framework import generics, status, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from django_filters.rest_framework import DjangoFilterBackend

from apps.users.permissions import IsAdmin, IsAdminOrReadOnly
from .models import Event, EventRegistration, JuryAssignment, EventStage, EventTask, RegistrationDocument
from .serializers import (
    EventListSerializer,
    EventDetailSerializer,
    EventCreateUpdateSerializer,
    EventRegistrationSerializer,
    JuryAssignmentSerializer,
    EventStageSerializer,
    EventTaskSerializer,
    RegistrationDocumentSerializer,
)


class EventViewSet(ModelViewSet):
    """
    Мероприятия (олимпиады и конкурсы).

    GET    /api/events/              — список (публично)
    POST   /api/events/              — создать (admin)
    GET    /api/events/{id}/         — детали (публично)
    PUT    /api/events/{id}/         — обновить (admin)
    DELETE /api/events/{id}/         — удалить (admin)
    POST   /api/events/{id}/register/    — зарегистрироваться (participant)
    POST   /api/events/{id}/unregister/  — отменить регистрацию (participant)
    GET    /api/events/{id}/participants/ — список участников (admin/teacher)
    GET    /api/events/{id}/stats/   — статистика (admin)
    POST   /api/events/{id}/assign_jury/ — назначить жюри (admin)
    """
    queryset = Event.objects.all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['event_type', 'movement_type', 'status', 'format']
    search_fields = ['title', 'description']
    ordering_fields = ['start_date', 'end_date', 'created_at']
    ordering = ['-start_date']

    def get_queryset(self):
        qs = Event.objects.all()
        if (self.request.query_params.get('my') == 'true'
                and self.request.user.is_authenticated):
            qs = qs.filter(registrations__participant=self.request.user)
        return qs

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticatedOrReadOnly()]
        if self.action in ('register', 'unregister'):
            return [IsAuthenticated()]
        if self.action == 'participants':
            return [IsAuthenticated()]
        return [IsAdmin()]

    def get_serializer_class(self):
        if self.action == 'list':
            return EventListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return EventCreateUpdateSerializer
        return EventDetailSerializer

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def register(self, request, pk=None):
        event = self.get_object()
        user = request.user

        if user.role != 'participant':
            return Response(
                {'detail': 'Только участники могут регистрироваться на мероприятия.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if not event.is_registration_open:
            return Response(
                {'detail': 'Регистрация на это мероприятие закрыта.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if EventRegistration.objects.filter(event=event, participant=user).exists():
            return Response(
                {'detail': 'Вы уже зарегистрированы на это мероприятие.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if event.max_participants and event.participant_count >= event.max_participants:
            return Response(
                {'detail': 'Достигнуто максимальное количество участников.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Определяем нужны ли документы
        needs_parental = event.requires_parental_consent or user.is_minor
        needs_application = event.requires_application
        needs_docs = needs_parental or needs_application

        reg_status = 'pending_docs' if needs_docs else 'registered'
        reg = EventRegistration.objects.create(event=event, participant=user, status=reg_status)

        from apps.notifications.models import Notification
        from apps.notifications.email_service import send_registration_email

        if needs_docs:
            required = []
            if needs_parental:
                required.append('согласие родителей/законного представителя')
            if needs_application:
                required.append('заявку на участие')
            Notification.objects.create(
                recipient=user,
                title=f'Регистрация на "{event.title}" — требуются документы',
                message=(
                    f'Для завершения регистрации необходимо загрузить: '
                    f'{", ".join(required)}. Перейдите в раздел «Документы» в своём кабинете.'
                ),
                notif_type='registration',
            )
        else:
            Notification.objects.create(
                recipient=user,
                title=f'Регистрация на "{event.title}"',
                message=f'Вы успешно зарегистрировались на мероприятие "{event.title}".',
                notif_type='registration',
            )
            send_registration_email(user, event)

        return Response(
            EventRegistrationSerializer(reg).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def unregister(self, request, pk=None):
        event = self.get_object()
        try:
            reg = EventRegistration.objects.get(event=event, participant=request.user)
        except EventRegistration.DoesNotExist:
            return Response(
                {'detail': 'Вы не зарегистрированы на это мероприятие.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        reg.status = 'withdrawn'
        reg.save(update_fields=['status'])
        return Response({'detail': 'Регистрация отменена.'})

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def participants(self, request, pk=None):
        event = self.get_object()
        user = request.user
        if user.role not in ('admin', 'teacher', 'jury'):
            return Response(status=status.HTTP_403_FORBIDDEN)
        registrations = event.registrations.select_related('participant').filter(
            status__in=('registered', 'active', 'completed')
        )
        serializer = EventRegistrationSerializer(registrations, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], permission_classes=[IsAdmin])
    def stats(self, request, pk=None):
        event = self.get_object()
        registrations = event.registrations
        submissions = event.submissions
        return Response({
            'total_registered': registrations.filter(
                status__in=('registered', 'active', 'completed')
            ).count(),
            'active': registrations.filter(status='active').count(),
            'completed': registrations.filter(status='completed').count(),
            'withdrawn': registrations.filter(status='withdrawn').count(),
            'submissions_total': submissions.count(),
            'submissions_evaluated': submissions.filter(status='evaluated').count(),
            'jury_count': event.jury_assignments.count(),
        })

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def assign_jury(self, request, pk=None):
        event = self.get_object()
        jury_id = request.data.get('jury_id')
        if jury_id and JuryAssignment.objects.filter(event=event, jury_id=jury_id).exists():
            return Response(
                {'detail': 'Этот эксперт уже назначен на данное мероприятие.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        serializer = JuryAssignmentSerializer(
            data={'event': event.id, **request.data}
        )
        serializer.is_valid(raise_exception=True)
        assignment = serializer.save()

        # Уведомление жюри
        from apps.notifications.models import Notification
        Notification.objects.create(
            recipient=assignment.jury,
            title=f'Назначение на "{event.title}"',
            message=f'Вас назначили жюри мероприятия "{event.title}".',
            notif_type='assignment',
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)


# ─── Registrations ────────────────────────────────────────────────────────────

class MyRegistrationsView(generics.ListAPIView):
    """GET /api/events/my-registrations/ — мои регистрации (для участника)."""
    permission_classes = [IsAuthenticated]
    serializer_class = EventRegistrationSerializer

    def get_queryset(self):
        return EventRegistration.objects.filter(
            participant=self.request.user
        ).select_related('event', 'participant')


# ─── Jury Assignments ─────────────────────────────────────────────────────────

class JuryAssignmentViewSet(ModelViewSet):
    """
    Назначения жюри.
    GET  /api/events/jury-assignments/      — список (admin видит все, жюри — свои)
    DELETE /api/events/jury-assignments/{id}/ — удалить (admin)
    """
    serializer_class = JuryAssignmentSerializer

    def get_permissions(self):
        if self.action in ('update', 'partial_update', 'create', 'destroy'):
            return [IsAdmin()]
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return JuryAssignment.objects.select_related('jury', 'event').all()
        return JuryAssignment.objects.filter(jury=user).select_related('jury', 'event')


# ─── Dashboard stats ──────────────────────────────────────────────────────────

class DashboardStatsView(generics.RetrieveAPIView):
    """GET /api/events/dashboard-stats/ — статистика для текущего пользователя."""
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        now = timezone.now().date()

        if user.role == 'admin':
            from django.contrib.auth import get_user_model
            User = get_user_model()
            return Response({
                'users_total': User.objects.count(),
                'events_active': Event.objects.filter(status='active').count(),
                'events_total': Event.objects.count(),
                'registrations_total': EventRegistration.objects.count(),
                'recent_registrations': EventRegistrationSerializer(
                    EventRegistration.objects.select_related('participant', 'event')
                    .order_by('-registered_at')[:5],
                    many=True,
                ).data,
            })

        if user.role == 'teacher':
            students = user.students.all()
            return Response({
                'students_count': students.count(),
                'active_events': Event.objects.filter(status='active').count(),
                'students_registered': EventRegistration.objects.filter(
                    participant__in=students
                ).values('participant').distinct().count(),
            })

        if user.role == 'participant':
            regs = EventRegistration.objects.filter(participant=user)
            from apps.submissions.models import EventResult
            return Response({
                'events_registered': regs.count(),
                'events_active': regs.filter(status='active').count(),
                'events_completed': regs.filter(status='completed').count(),
                'awards': EventResult.objects.filter(participant=user, place__gt=0).count(),
                'upcoming_deadlines': Event.objects.filter(
                    registrations__participant=user,
                    registration_deadline__gte=now,
                    status__in=('upcoming', 'active'),
                ).count(),
            })

        if user.role == 'jury':
            from apps.submissions.models import Submission, Evaluation
            assigned = Submission.objects.filter(jury=user)
            return Response({
                'assigned_total': assigned.count(),
                'evaluated': assigned.filter(status='evaluated').count(),
                'pending': assigned.filter(status__in=('submitted', 'under_review')).count(),
                'events_count': JuryAssignment.objects.filter(jury=user).count(),
            })

        return Response({})


# ─── Analytics ────────────────────────────────────────────────────────────────

class GlobalAnalyticsView(generics.GenericAPIView):
    """GET /api/events/analytics/?year=2025&event_id=1 — сводная аналитика (admin/teacher)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        if user.role not in ('admin', 'teacher'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        year = request.query_params.get('year')
        event_id = request.query_params.get('event_id')

        from apps.submissions.models import Submission
        from django.db.models.functions import TruncMonth
        from django.db.models import Avg

        if user.role == 'admin':
            events_qs = Event.objects.all()
            subs_qs = Submission.objects.all()
            regs_qs = EventRegistration.objects.all()
        else:  # teacher
            student_ids = list(user.students.values_list('id', flat=True))
            events_qs = Event.objects.filter(registrations__participant__in=student_ids).distinct()
            subs_qs = Submission.objects.filter(participant__in=student_ids)
            regs_qs = EventRegistration.objects.filter(participant__in=student_ids)

        if year:
            events_qs = events_qs.filter(start_date__year=year)
            subs_qs = subs_qs.filter(created_at__year=year)
            regs_qs = regs_qs.filter(registered_at__year=year)

        if event_id:
            subs_qs = subs_qs.filter(event_id=event_id)
            regs_qs = regs_qs.filter(event_id=event_id)

        monthly = list(
            subs_qs.exclude(status='draft')
            .filter(submitted_at__isnull=False)
            .annotate(month=TruncMonth('submitted_at'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )

        per_event = []
        for ev in events_qs.order_by('-start_date')[:30]:
            ev_subs = subs_qs.filter(event=ev)
            ev_regs = regs_qs.filter(event=ev)
            per_event.append({
                'id': ev.id,
                'title': ev.title,
                'status': ev.status,
                'start_date': str(ev.start_date) if ev.start_date else None,
                'registrations': ev_regs.count(),
                'submissions': ev_subs.exclude(status='draft').count(),
                'evaluated': ev_subs.filter(status='evaluated').count(),
            })

        return Response({
            'summary': {
                'events_total': events_qs.count(),
                'registrations_total': regs_qs.count(),
                'submissions_total': subs_qs.exclude(status='draft').count(),
                'evaluated_total': subs_qs.filter(status='evaluated').count(),
            },
            'monthly_chart': [
                {'month': m['month'].strftime('%Y-%m'), 'count': m['count']}
                for m in monthly if m['month']
            ],
            'per_event': per_event,
        })


class EventDetailAnalyticsView(generics.GenericAPIView):
    """GET /api/events/{id}/analytics/ — детальная аналитика по мероприятию."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk, *args, **kwargs):
        from apps.submissions.models import Submission, Evaluation
        from django.db.models.functions import TruncDate
        from django.db.models import Avg

        user = request.user
        if user.role not in ('admin', 'teacher'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        try:
            event = Event.objects.get(pk=pk)
        except Event.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        subs_qs = Submission.objects.filter(event=event)
        regs_qs = EventRegistration.objects.filter(event=event)
        evals_qs = Evaluation.objects.filter(submission__event=event, is_draft=False)

        if user.role == 'teacher':
            student_ids = list(user.students.values_list('id', flat=True))
            subs_qs = subs_qs.filter(participant__in=student_ids)
            regs_qs = regs_qs.filter(participant__in=student_ids)
            evals_qs = evals_qs.filter(submission__participant__in=student_ids)

        avg_score = evals_qs.aggregate(avg=Avg('score'))['avg']

        score_ranges = [
            {'range': '0–20', 'count': evals_qs.filter(score__lte=20).count()},
            {'range': '21–40', 'count': evals_qs.filter(score__gt=20, score__lte=40).count()},
            {'range': '41–60', 'count': evals_qs.filter(score__gt=40, score__lte=60).count()},
            {'range': '61–80', 'count': evals_qs.filter(score__gt=60, score__lte=80).count()},
            {'range': '81–100', 'count': evals_qs.filter(score__gt=80).count()},
        ]

        daily = list(
            subs_qs.exclude(status='draft')
            .filter(submitted_at__isnull=False)
            .annotate(day=TruncDate('submitted_at'))
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )

        stages = EventStage.objects.filter(event=event).prefetch_related('tasks').order_by('order')
        stages_data = [
            {'id': s.id, 'title': s.title, 'order': s.order, 'tasks_count': s.tasks.count()}
            for s in stages
        ]

        return Response({
            'event': {
                'id': event.id, 'title': event.title, 'status': event.status,
                'start_date': str(event.start_date) if event.start_date else None,
                'end_date': str(event.end_date) if event.end_date else None,
            },
            'summary': {
                'registrations': regs_qs.count(),
                'submissions': subs_qs.exclude(status='draft').count(),
                'evaluated': subs_qs.filter(status='evaluated').count(),
                'avg_score': round(avg_score, 1) if avg_score else None,
            },
            'score_distribution': score_ranges,
            'submissions_timeline': [
                {'date': d['day'].strftime('%Y-%m-%d'), 'count': d['count']}
                for d in daily if d['day']
            ],
            'stages': stages_data,
        })


class AnalyticsExportView(generics.GenericAPIView):
    """GET /api/events/analytics/export/ — выгрузка отчёта в Excel."""
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        user = request.user
        if user.role not in ('admin', 'teacher'):
            return Response(status=status.HTTP_403_FORBIDDEN)

        year = request.query_params.get('year')
        event_id = request.query_params.get('event_id')
        report_type = request.query_params.get('type', 'submissions')

        from apps.submissions.models import Submission, Evaluation
        from django.db.models import Avg

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1E3A5F')
        center = Alignment(horizontal='center')

        def style_headers(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

        def auto_width(ws):
            for col in ws.columns:
                width = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(width + 4, 50)

        STATUS_LABELS = {
            'draft': 'Черновик', 'submitted': 'Отправлено',
            'under_review': 'На проверке', 'evaluated': 'Проверено',
        }
        EVENT_STATUS = {
            'draft': 'Черновик', 'upcoming': 'Предстоящее',
            'active': 'Активное', 'completed': 'Завершено', 'cancelled': 'Отменено',
        }
        EVENT_TYPE = {'olympiad': 'Олимпиада', 'competition': 'Конкурс'}

        if report_type == 'submissions':
            ws = wb.active
            ws.title = 'Работы участников'
            headers = ['ID', 'Участник', 'Email', 'Мероприятие', 'Статус', 'Дата подачи', 'Оценка', 'Комментарий жюри']
            style_headers(ws, headers)

            subs_qs = (Submission.objects
                       .select_related('participant', 'event')
                       .exclude(status='draft'))
            if user.role == 'teacher':
                subs_qs = subs_qs.filter(participant__in=user.students.all())
            if year:
                subs_qs = subs_qs.filter(submitted_at__year=year)
            if event_id:
                subs_qs = subs_qs.filter(event_id=event_id)

            for sub in subs_qs:
                try:
                    ev_obj = sub.evaluation
                    score = ev_obj.score if not ev_obj.is_draft else ''
                    feedback = ev_obj.feedback if not ev_obj.is_draft else ''
                except Exception:
                    score, feedback = '', ''
                ws.append([
                    sub.id,
                    sub.participant.full_name or f'{sub.participant.last_name} {sub.participant.first_name}'.strip(),
                    sub.participant.email,
                    sub.event.title,
                    STATUS_LABELS.get(sub.status, sub.status),
                    sub.submitted_at.strftime('%d.%m.%Y %H:%M') if sub.submitted_at else '',
                    score,
                    feedback,
                ])
            auto_width(ws)

        elif report_type == 'events':
            ws = wb.active
            ws.title = 'Мероприятия'
            headers = ['ID', 'Название', 'Тип', 'Статус', 'Начало', 'Конец', 'Регистраций', 'Работ подано', 'Проверено', 'Ср. балл']
            style_headers(ws, headers)

            events_qs = Event.objects.all()
            if user.role == 'teacher':
                student_ids = list(user.students.values_list('id', flat=True))
                events_qs = Event.objects.filter(registrations__participant__in=student_ids).distinct()
            if year:
                events_qs = events_qs.filter(start_date__year=year)
            if event_id:
                events_qs = events_qs.filter(id=event_id)

            for ev in events_qs.order_by('-start_date'):
                subs = Submission.objects.filter(event=ev).exclude(status='draft')
                regs = EventRegistration.objects.filter(event=ev)
                if user.role == 'teacher':
                    student_ids = list(user.students.values_list('id', flat=True))
                    subs = subs.filter(participant__in=student_ids)
                    regs = regs.filter(participant__in=student_ids)
                avg = Evaluation.objects.filter(submission__event=ev, is_draft=False).aggregate(a=Avg('score'))['a']
                ws.append([
                    ev.id, ev.title,
                    EVENT_TYPE.get(ev.event_type, ev.event_type),
                    EVENT_STATUS.get(ev.status, ev.status),
                    ev.start_date.strftime('%d.%m.%Y') if ev.start_date else '',
                    ev.end_date.strftime('%d.%m.%Y') if ev.end_date else '',
                    regs.count(), subs.count(),
                    subs.filter(status='evaluated').count(),
                    round(avg, 1) if avg else '',
                ])
            auto_width(ws)

        elif report_type == 'participants':
            ws = wb.active
            ws.title = 'Участники'
            headers = ['ID', 'ФИО', 'Email', 'Учреждение', 'Мероприятий', 'Работ подано', 'Оценок получено', 'Ср. балл']
            style_headers(ws, headers)

            from django.contrib.auth import get_user_model
            User_model = get_user_model()
            participants_qs = User_model.objects.filter(role='participant')
            if user.role == 'teacher':
                participants_qs = user.students.all()
            if year:
                participants_qs = participants_qs.filter(
                    registrations__registered_at__year=year
                ).distinct()

            for p in participants_qs:
                regs = EventRegistration.objects.filter(participant=p)
                subs = Submission.objects.filter(participant=p).exclude(status='draft')
                if year:
                    subs = subs.filter(submitted_at__year=year)
                evals = Evaluation.objects.filter(submission__participant=p, is_draft=False)
                avg = evals.aggregate(a=Avg('score'))['a']
                ws.append([
                    p.id,
                    p.full_name or f'{p.last_name} {p.first_name}'.strip(),
                    p.email,
                    p.institution or '',
                    regs.count(), subs.count(), evals.count(),
                    round(avg, 1) if avg else '',
                ])
            auto_width(ws)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'report_{report_type}{"_" + year if year else ""}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ─── Stages ───────────────────────────────────────────────────────────────────

class EventStageViewSet(ModelViewSet):
    """
    Этапы мероприятий.
    GET    /api/events/stages/?event=<id>   — список этапов мероприятия
    POST   /api/events/stages/              — создать этап (admin)
    PUT    /api/events/stages/{id}/         — обновить (admin)
    DELETE /api/events/stages/{id}/         — удалить (admin)
    """
    serializer_class = EventStageSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['event']

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsAdmin()]

    def get_queryset(self):
        return EventStage.objects.prefetch_related('tasks').all()


# ─── Tasks ────────────────────────────────────────────────────────────────────

class EventTaskViewSet(ModelViewSet):
    """
    Задания этапов.
    GET    /api/events/tasks/?stage=<id>   — список заданий этапа
    POST   /api/events/tasks/              — создать задание (admin)
    PUT    /api/events/tasks/{id}/         — обновить (admin)
    DELETE /api/events/tasks/{id}/         — удалить (admin)
    """
    serializer_class = EventTaskSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['stage']

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsAdmin()]

    def get_queryset(self):
        return EventTask.objects.all()

    def perform_destroy(self, instance):
        if instance.file:
            instance.file.delete(save=False)
        instance.delete()


# ─── Registration Documents ────────────────────────────────────────────────────

class RegistrationDocumentViewSet(ModelViewSet):
    """
    Документы для регистрации (согласия, заявки).

    Участник:
      POST /api/events/reg-docs/         — загрузить документ
      GET  /api/events/reg-docs/         — мои документы
      GET  /api/events/reg-docs/pending/ — регистрации, ожидающие документов (admin)

    Администратор:
      POST /api/events/reg-docs/{id}/approve/ — одобрить документ
      POST /api/events/reg-docs/{id}/reject/  — отклонить документ
    """
    serializer_class = RegistrationDocumentSerializer

    def get_permissions(self):
        if self.action in ('approve', 'reject', 'pending_registrations'):
            return [IsAdmin()]
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return RegistrationDocument.objects.select_related(
                'registration__participant', 'registration__event'
            ).all()
        return RegistrationDocument.objects.filter(
            registration__participant=user
        ).select_related('registration__event')

    def perform_create(self, serializer):
        # Проверяем, что регистрация принадлежит текущему пользователю
        reg = serializer.validated_data['registration']
        if reg.participant != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Эта регистрация вам не принадлежит.')
        serializer.save()

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def approve(self, request, pk=None):
        doc = self.get_object()
        doc.status = 'approved'
        doc.admin_note = request.data.get('admin_note', '')
        doc.reviewed_by = request.user
        doc.reviewed_at = timezone.now()
        doc.save(update_fields=['status', 'admin_note', 'reviewed_by', 'reviewed_at'])

        # Если все документы регистрации одобрены — активируем регистрацию
        reg = doc.registration
        all_docs = reg.documents.all()
        if all_docs.exists() and not all_docs.filter(status__in=('pending', 'rejected')).exists():
            reg.status = 'registered'
            reg.save(update_fields=['status'])
            from apps.notifications.models import Notification
            from apps.notifications.email_service import send_registration_email
            Notification.objects.create(
                recipient=reg.participant,
                title=f'Регистрация на "{reg.event.title}" подтверждена',
                message='Все ваши документы проверены и одобрены. Регистрация завершена!',
                notif_type='registration',
            )
            send_registration_email(reg.participant, reg.event)

        return Response({'detail': 'Документ одобрен.'})

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def reject(self, request, pk=None):
        doc = self.get_object()
        doc.status = 'rejected'
        doc.admin_note = request.data.get('admin_note', '')
        doc.reviewed_by = request.user
        doc.reviewed_at = timezone.now()
        doc.save(update_fields=['status', 'admin_note', 'reviewed_by', 'reviewed_at'])

        from apps.notifications.models import Notification
        reg = doc.registration
        Notification.objects.create(
            recipient=reg.participant,
            title=f'Документ для "{reg.event.title}" отклонён',
            message=(
                f'Ваш документ «{doc.get_doc_type_display()}» отклонён. '
                f'{doc.admin_note or "Загрузите исправленный документ."}'
            ),
            notif_type='registration',
        )
        return Response({'detail': 'Документ отклонён.'})

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def pending_registrations(self, request):
        """GET /api/events/reg-docs/pending_registrations/ — все регистрации ожидающие проверки."""
        regs = EventRegistration.objects.filter(
            status='pending_docs'
        ).select_related('participant', 'event').prefetch_related('documents')
        return Response(EventRegistrationSerializer(regs, many=True).data)
